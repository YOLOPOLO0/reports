import openpyxl
from openpyxl.styles import protection, Protection
from openpyxl import Workbook, load_workbook 
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import  Border, Side, BORDER_THIN
wb = Workbook()
ws = wb.active

ws.title = "CWC"
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 15.57
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 19.57
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

set_border(ws, 'A3:D3') 
font = openpyxl.styles.Font(bold = True , size = 16)
font2 = openpyxl.styles.Font(bold = True , size = 11, underline='single')
ws['A1'].font = font
ws['A1'].alignment = Alignment(horizontal='center')
ws.merge_cells('A1:D1')
ws['A1']= 'Corrugated Box'
ws['A2'].font = font
ws['A2'].alignment = Alignment(horizontal='center')
ws.merge_cells('A2:D2')
ws['A2']= 'Certificate of Analysis'
ws['A3'] = 'Variant/Product:'
ws.merge_cells('B3:D3')
ws['B3']='Product*Grammage*Price (no. of packs/box)'
ws['B3'].alignment = Alignment(horizontal='center')
ws.append(['','','',''])
ws.append(['Date:','00-Jan-3000','Traceability #'])
ws.append(['Batch #','XXXX','Quantity:','XXXX'])
ws.append(['Date of Maanufacturing','00-Jan-3000','Date of Dispatch','00-Jan-3000'])
ws.append(['Date of Expiry','00-Jan-3000',])
set_border(ws, 'A5:D8')
ws.append(['','','',''])
ws.append(['Box Type:','                              '])
ws.append(['Paper Description:','xxx','xxx','xxx'])
ws.append(['Flute Type','','Flute Direction:',''])
set_border(ws, 'A10:D12')
ws.append(['','','',''])
ws['A14']= 'Parameter'
ws['A14'].alignment = Alignment(horizontal='center')
ws['A14'].font = font2
ws['B14']= 'UOM'
ws['B14'].alignment = Alignment(horizontal='center')
ws['B14'].font = font2
ws['C14']= 'Specification'
ws['C14'].alignment = Alignment(horizontal='center')
ws['C14'].font = font2
# ws['D14']= 'Results'
# ws['D14'].alignment = Alignment(horizontal='center')
# ws['D14'].font = font2
ws.append(['Moisture','%'])
ws.append(['Internal Case Length','(mm)'])
ws.append(['Internal Case Width','(mm)'])
ws.append(['Internal Case Height','(mm)'])
ws.append(['Tape Length','(mm)'])
ws.append(['Specification Type'])
ws.append(['Carton Type','Regular/Strong'])
ws.append(['Bags/Case','no.'])
ws.append(['Orientation','H,V'])
ws.append(['Config(rows x col x layers)'])
ws.append(['Grammage','g/m2'])
ws.append(['Thickness','(mm)'])
ws.append(['Flat Crush Strength','KPa'])
ws.append(['Edge Crush Strength','KN/m'])
ws.append(['Box Compression Strength','kgf'])
ws.append(['Thickness','(mm)'])
set_border(ws, 'A14:D30')
ws.append(['','','',''])
ws.append(['Gap B/W the Flaps (Top&Bottom)','(mm)','<3mm'])
ws.append(['Glued/Stitched'])
ws.append(['MFG. Joint (Inner/Outter)'])
ws.append(['Printed/Unprinted'])
ws.append(['Color'])
set_border(ws, 'A32:D36')
ws.merge_cells('A38:D39')
ws['A38']='Note: We certify that the supplied product has been inspected tested and free of infestation in accordance with ISO 9001 Quality system and hence conforms to agreed standards.'
ws['A38'].alignment = Alignment(horizontal='justify')
ws['A43']='Inspected By.'
ws['A43'].font = font2
ws['D43']='Quality Controller'

ws.protection.sheet = True


cell = ws['A1:D43']
for i in cell:
    for a in i:
        if not a.value:
            a.protection = Protection(locked=False)






wb.save('lool.xlsx')
